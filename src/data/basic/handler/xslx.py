from typing import Type, cast

import openpyxl
from openpyxl.workbook import Workbook

from data.basic.generator import generate_people
from data.basic.model_dataclasses import Person, Car, Airport


def write_people(people: list[Person], wb: Workbook,
                 sheet_name: str = "people",
                 heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["id", "name", "age", "male"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(people)):
        sheet.cell(row=row + offset, column=1, value=people[row].id)
        sheet.cell(row=row + offset, column=2, value=people[row].name)
        sheet.cell(row=row + offset, column=3, value=people[row].age)
        sheet.cell(row=row + offset, column=4, value=people[row].male)


def read_people(wb: Workbook,
                sheet_name: str = "people",
                heading: bool = True) -> list[Person]:
    sheet = wb[sheet_name]

    people = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        people.append(
            Person(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value
            )
        )
        row += 1
    return people

def read_cars(workbook: openpyxl.Workbook, sheet_name: str = "cars", heading: bool = True) -> list[Car]:
    sheet = workbook[sheet_name if sheet_name is not None else "cars"]
    cars = []

    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        cars.append(Car(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            int(sheet.cell(row=row, column=3).value),
            bool(sheet.cell(row=row, column=4).value)
        ))
        row += 1

    return cars


def write_cars(cars: list[Car], workbook: openpyxl.Workbook, sheet_name: str = "cars", heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "cars")

    if heading:
        field_names = ["plate", "type", "year", "automatic"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(cars)):
        sheet.cell(row=row + offset, column=1, value=cars[row].plate)
        sheet.cell(row=row + offset, column=2, value=cars[row].type)
        sheet.cell(row=row + offset, column=3, value=cars[row].year)
        sheet.cell(row=row + offset, column=4, value=cars[row].automatic)


def read_airports(workbook: openpyxl.Workbook, sheet_name: str = "airports", heading: bool = True) -> list[Airport]:
    sheet = workbook[sheet_name if sheet_name is not None else "airports"]
    airports = []

    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        airports.append(Airport(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
            sheet.cell(row=row, column=4).value,
            sheet.cell(row=row, column=4).value
        ))
        row += 1

    return airports


def write_airports(airports: list[Airport], workbook: openpyxl.Workbook, sheet_name: str = "airports",
                   heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "airports")

    if heading:
        field_names = ["code", "name", "city", "state", "country"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(airports)):
        sheet.cell(row=row + offset, column=1, value=airports[row].code)
        sheet.cell(row=row + offset, column=2, value=airports[row].name)
        sheet.cell(row=row + offset, column=3, value=airports[row].city)
        sheet.cell(row=row + offset, column=4, value=airports[row].state)
        sheet.cell(row=row + offset, column=5, value=airports[row].country)


def read(entity_type: Type[object], workbook: openpyxl.Workbook,
         sheet_name: str = None, heading: bool = True) -> list[object]:
    if entity_type == Person:
        return read_people(workbook, sheet_name=sheet_name, heading=heading)
    elif entity_type == Car:
        return read_cars(workbook, sheet_name=sheet_name, heading=heading)
    elif entity_type == Airport:
        return read_airports(workbook, sheet_name=sheet_name, heading=heading)
    else:
        raise RuntimeError("Unknown type of entity")


def write(entities: list[object], workbook: openpyxl.Workbook,
          sheet_name: str = None, heading: bool = True) -> None:
    if isinstance(entities[0], Person):
        return write_people([cast(Person, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
    elif isinstance(entities[0], Car):
        return write_cars([cast(Car, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
    elif isinstance(entities[0], Airport):
        return write_airports([cast(Airport, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
    else:
        raise RuntimeError("Unknown type of entity")


if __name__ == "__main__":
    wb = Workbook()
    write_people(
        generate_people(10),
        wb
    )
    wb.save("D:/people.xlsx")

    wb = openpyxl.load_workbook("D:/people.xlsx")
    for person in read_people(wb):
        print(person)


